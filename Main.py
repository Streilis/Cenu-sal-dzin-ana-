from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import openpyxl
import re

#Excel dokumenta izveidošana un sagatavošana
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet2 = workbook.create_sheet('Cenu salīdzinājums')
sheet2['A2'] = '1a.lv'
sheet2['A3'] = 'Rdveikals.lv'
sheet2['A4'] = '220.lv'
sheet2['B1'] = 'Lētākie modeļi'
sheet2['C1'] = 'Cena'
sheet2.column_dimensions['B'].width = 59
sheet2.column_dimensions['C'].width = 16



sheet.title = 'Visi dati'
sheet.merge_cells('A1:C1')
sheet.merge_cells('D1:E1')
sheet.merge_cells('F1:G1')
for col in ['A1', 'D1','F1','A2','B2','C2','D2','E2','F2','G2']:
    sheet[col].alignment = openpyxl.styles.Alignment(horizontal='center')
for col in ['A','D','F']:
    sheet.column_dimensions[col].width = 59
for col in ['B','C','E','G']:
    sheet.column_dimensions[col].width = 16
for col in ['C2','E2','G2']:
    sheet[col] = 'Cena'
for col in ['A2','D2','F2']:
    sheet[col] = 'Modelis'  
sheet['A1'] = '1a.lv'
sheet['D1'] = 'Rdveikals.lv'
sheet['F1'] = '220.lv'
sheet['B2'] = 'Cena ar Klienta karti'



first_market = 'https://www.1a.lv/c/telefoni-plansetdatori-apple-veikals/mobilie-telefoni-viedtalruni/iphone-apple-telefoni/asi'
all_products_1 = '//*[@id="catalog-page"]/div[3]/div[3]/div/div'
product_name_1 = './/div[2]/a[2]'
price_with_discount = './/div/div/div[2]/div/span[1]'
product_price_1 = './/div[2]/div[1]/div/span/span[1]'
first_phone_model = "Mobilais telefons Apple iPhone 16 Pro Max"
cokies2 = '//*[@id="CybotCookiebotDialogBodyLevelButtonLevelOptinAllowAll"]'

second_market = "https://www.rdveikals.lv/categories/lv/388/sort/5/filter/0_0_0_201300/page/1/Mobilie-telefoni.html"
all_products_2 ='//*[@id="main_container_wrapper"]/ul[1]/li | //*[@id="main_container_wrapper"]/ul[2]/li'
product_name_2 ='.//div[3]/div[1]/h3/a'
product_price_2 = './/div[3]/p'


third_market = "https://220.lv/lv/mobilie-telefoni/f/apple/apple-iphone-16-pro-max"
all_products3 = '/html/body/div[7]/div/section[2]/div/div/div[2]/section/div[7]/div/div'
product_name3 = './/div[4]/a | .//div[5]/a'
product_price3 = './/div[6]/span | .//div[7]/span[1]'
cokies_3 = '//*[@id="consent_block"]/div/div[2]/div[2]/button[1]'
out_of_store = './/div[3]/span/span | .//[4]/span/span'
phone_model = "iPhone 16 Pro Max"

driver = webdriver.Chrome()  

def open_page_collect_info(market_url, cokies, all_data):
    time.sleep(3)
    driver.get(market_url)
    time.sleep(3)
    try:
        driver.find_element(By.XPATH, cokies).click()
        time.sleep(3)
    except:
        print("Nav cookies.")
    return driver.find_elements(By.XPATH, all_data)

def cenas_formats(cena):
    cena = cena.replace(' €', '')
    cena = cena.replace('.', '')
    cena = cena.replace(',', '.')
    return float(cena)


def nolasit_produktus_1(row=3):
    letaka_cena = None
    letakais_modelis = None
    all_data = open_page_collect_info(first_market, cokies2, all_products_1)
    for produkts in all_data:
        try:
            nosaukums = produkts.find_element(By.XPATH, product_name_1).text
            if phone_model in nosaukums:
                try:
                    cena_ar_atlaidi = produkts.find_element(By.XPATH, price_with_discount).text
                    cena_ar_atlaidi = cenas_formats(cena_ar_atlaidi)
                except:
                    cena_ar_atlaidi = 'Nav atlaides'
                cena = produkts.find_element(By.XPATH, product_price_1).text
                cena = cenas_formats(cena)
                if cena_ar_atlaidi == 'Nav atlaides':
                    if letaka_cena is None or cena < letaka_cena:
                        letaka_cena = cena
                        letakais_modelis = nosaukums
                elif letaka_cena is None or cena_ar_atlaidi < letaka_cena:
                    letaka_cena = cena_ar_atlaidi
                    letakais_modelis = nosaukums
                sheet[f"A{row}"] = nosaukums
                sheet[f"B{row}"] = f"{cena_ar_atlaidi:.2f} €"
                sheet[f"C{row}"] = f"{cena:.2f} €"
                row += 1    
            else:
                continue      
        except: 
            pass
    sheet2["B2"] = letakais_modelis
    sheet2["C2"] = f"{letaka_cena:.2f} €"
    

def nolasit_produktus_2(row=3):
    letaka_cena = None
    letakais_modelis = None
    all_data = open_page_collect_info(second_market, cokies2, all_products_2)
    for produkts in all_data:
        try:
            nosaukums = produkts.find_element(By.XPATH, product_name_2).text
            if phone_model in nosaukums:
                cena = produkts.find_element(By.XPATH, product_price_2).text
                cena = cena.replace(' €', '')
                cena = float(cena)
                if letaka_cena is None or cena < letaka_cena:
                    letaka_cena = cena
                    letakais_modelis = nosaukums
                sheet[f"D{row}"] = nosaukums
                sheet[f"E{row}"] = f"{cena:.2f} €"
                row += 1
            else:
                continue
        except:
            pass
    sheet2["B3"] = letakais_modelis
    sheet2["C3"] = f"{letaka_cena:.2f} €"



def nolasit_produktus_3(row=3):
    
    letaka_cena = None
    letakais_modelis = None
    all_data = open_page_collect_info(third_market, cokies_3, all_products3)
    for produkts in all_data:
        try:
            nosaukums_elementa = produkts.find_element(By.XPATH, product_name3)
            nosaukums = nosaukums_elementa.text
            
            if phone_model in nosaukums:  
                visas_cenas = produkts.find_elements(By.XPATH, product_price3)
                pilna_cena = []
                for i in visas_cenas:
                    if re.match(r"^\(\d+\)$",i.text) or re.match(r"^\d+\.\d+$",i.text):
                        continue
                    else:
                        pilna_cena.append(i.text.strip())
                cena_full = pilna_cena[0].split('\n')
                cena = cena_full[0] +"."+ cena_full[1]
                cena = cena.replace(" ","")
                cena = float(cena)
                if letaka_cena is None or cena < letaka_cena:
                    letaka_cena = cena
                    letakais_modelis = nosaukums
                sheet[f"G{row}"] = f"{cena:.2f} €"
                sheet[f"F{row}"] = nosaukums
                row += 1    
            else:
                continue      
        except:
            pass
    sheet2["B4"] = letakais_modelis
    sheet2["C4"] = f"{letaka_cena:.2f} €"


def main():
    
    nolasit_produktus_1()
    nolasit_produktus_2()
    nolasit_produktus_3()
    driver.quit()
    workbook.save('Cenu_salidzinasana.xlsx')

if __name__ == "__main__":
    main()






